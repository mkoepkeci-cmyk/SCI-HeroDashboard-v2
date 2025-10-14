"""
Analyze Data Quality from Individual SCI Sheets

This script reads each team member's individual sheet and analyzes:
1. How many assignments have work effort defined
2. How many are missing work effort
3. Data completion rate
4. Work type breakdown with hours

Output: JSON file with data quality metrics for dashboard
"""

import openpyxl
import json
import sys

# Excel file path
EXCEL_FILE = 'documents/SCI Workload Tracker - New System.xlsx'

# Team member sheets to analyze
TEAM_MEMBERS = [
    'Ashley', 'Brooke', 'Dawn', 'Jason', 'Josh', 'Kim', 'Lisa',
    'Marisa', 'Marty', 'Matt', 'Melissa', 'Robin', 'Sherry',
    'Trudy', 'Van', 'Yvette'
]

def parse_work_effort_hours(effort_str):
    """Parse work effort string to get hour estimate"""
    if not effort_str:
        return None

    effort_map = {
        'XS': 0.5,
        'S': 1.5,
        'M': 3.5,
        'L': 7.5,
        'XL': 15.0
    }

    effort_str = str(effort_str).upper().strip()

    for key, hours in effort_map.items():
        if key in effort_str:
            return hours

    return None

def analyze_individual_sheet(worksheet, member_name):
    """Analyze a single team member's sheet"""

    result = {
        'name': member_name,
        'total_assignments': 0,
        'has_work_effort': 0,
        'missing_work_effort': 0,
        'completion_rate': 0.0,
        'active_assignments': 0,
        'estimated_hours': 0.0,
        'work_types': {},
        'missing_by_work_type': {},
        'assignments_detail': []
    }

    # Read assignments starting from row 2 (row 1 is header)
    for row_idx in range(2, worksheet.max_row + 1):
        row = worksheet[row_idx]

        # Column A: SCI name
        # Column B: Assignment name
        # Column E: Work Effort
        # Column G: Work Type
        # Column I: Status

        sci_name = row[0].value
        assignment_name = row[1].value
        work_effort = row[4].value
        work_type = row[6].value
        status = row[8].value

        # Check if this is an assignment row
        if not sci_name or not assignment_name:
            continue

        # Parse work effort hours
        effort_hours = parse_work_effort_hours(work_effort)
        has_effort = effort_hours is not None

        # Check if active
        is_active = status and ('Progress' in str(status) or 'Planning' in str(status))

        # Update counts
        result['total_assignments'] += 1
        if has_effort:
            result['has_work_effort'] += 1
        else:
            result['missing_work_effort'] += 1

        if is_active:
            result['active_assignments'] += 1
            if effort_hours:
                result['estimated_hours'] += effort_hours

        # Track work types
        if work_type:
            wt = str(work_type).strip()
            if wt not in result['work_types']:
                result['work_types'][wt] = {'count': 0, 'hours': 0, 'missing': 0}

            result['work_types'][wt]['count'] += 1
            if effort_hours and is_active:
                result['work_types'][wt]['hours'] += effort_hours
            if not has_effort:
                result['work_types'][wt]['missing'] += 1

        # Store assignment detail
        result['assignments_detail'].append({
            'name': str(assignment_name)[:50],
            'work_effort': str(work_effort) if work_effort else 'MISSING',
            'work_type': str(work_type) if work_type else 'Unknown',
            'status': str(status) if status else 'Unknown',
            'has_effort': has_effort,
            'is_active': is_active,
            'estimated_hours': effort_hours
        })

    # Calculate completion rate
    if result['total_assignments'] > 0:
        result['completion_rate'] = result['has_work_effort'] / result['total_assignments']

    return result

def main():
    print('Analyzing Data Quality from Excel Sheets')
    print('=' * 70)

    # Load workbook
    print(f'Reading: {EXCEL_FILE}')
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    all_results = []

    # Analyze each team member
    for member_name in TEAM_MEMBERS:
        if member_name not in wb.sheetnames:
            print(f'  Sheet not found: {member_name}')
            continue

        ws = wb[member_name]
        result = analyze_individual_sheet(ws, member_name)

        all_results.append(result)

        # Print summary
        print(f'\n{member_name}:')
        print(f'  Total: {result["total_assignments"]:3d} assignments')
        print(f'  Active: {result["active_assignments"]:3d} assignments')
        print(f'  Has Effort: {result["has_work_effort"]:3d} ({result["completion_rate"]*100:.1f}%)')
        print(f'  Missing: {result["missing_work_effort"]:3d}')
        print(f'  Est. Hours/Week: {result["estimated_hours"]:.1f}h')

        if result['work_types']:
            print(f'  Work Types:')
            for wt, data in sorted(result['work_types'].items(), key=lambda x: x[1]['hours'], reverse=True):
                if data['count'] > 0:
                    print(f'    {wt:25s}: {data["count"]:2d} items, {data["hours"]:5.1f}h, {data["missing"]} missing')

    # Calculate team-wide stats
    print('\n' + '=' * 70)
    print('TEAM SUMMARY:')
    print('=' * 70)

    total_assignments = sum(r['total_assignments'] for r in all_results)
    total_has_effort = sum(r['has_work_effort'] for r in all_results)
    total_missing = sum(r['missing_work_effort'] for r in all_results)
    total_active_hours = sum(r['estimated_hours'] for r in all_results)
    avg_completion = sum(r['completion_rate'] for r in all_results) / len(all_results)

    print(f'Total Assignments: {total_assignments}')
    print(f'Has Work Effort: {total_has_effort} ({(total_has_effort/total_assignments)*100:.1f}%)')
    print(f'Missing Work Effort: {total_missing} ({(total_missing/total_assignments)*100:.1f}%)')
    print(f'Total Active Hours/Week: {total_active_hours:.1f}h')
    print(f'Average Completion Rate: {avg_completion*100:.1f}%')

    # Find members with worst data quality
    print('\nMEMBERS NEEDING DATA UPDATES (worst data quality):')
    sorted_by_quality = sorted(all_results, key=lambda x: x['completion_rate'])
    for r in sorted_by_quality[:5]:
        print(f'  {r["name"]:12s}: {r["completion_rate"]*100:5.1f}% complete ({r["missing_work_effort"]} missing)')

    # Save to JSON
    output_file = 'data-quality-analysis.json'
    with open(output_file, 'w') as f:
        json.dump({
            'team_members': all_results,
            'team_summary': {
                'total_assignments': total_assignments,
                'has_effort': total_has_effort,
                'missing_effort': total_missing,
                'total_active_hours': total_active_hours,
                'avg_completion_rate': avg_completion
            }
        }, f, indent=2)

    print(f'\nAnalysis saved to: {output_file}')

    wb.close()

if __name__ == '__main__':
    main()
