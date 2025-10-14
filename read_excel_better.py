import openpyxl
import sys

# Force UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('documents/SCI Workload Tracker - New System.xlsx', data_only=True)

# Print all sheet names
print("=" * 80)
print("SHEET NAMES:")
print("=" * 80)
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Focus on individual team member sheets
team_member_sheets = [
    'Ashley', 'Brooke', 'Dawn', 'Jason', 'Josh', 'Kim', 'Lisa',
    'Marisa', 'Marty', 'Matt', 'Melissa', 'Robin', 'Sherry',
    'Trudy', 'Van', 'Yvette', 'Tiffany', 'Carrie'
]

print("\n" + "=" * 80)
print("TEAM MEMBER SHEET STRUCTURE (showing first 3 members):")
print("=" * 80)

# Show detailed structure for first 3 members
for sheet_name in team_member_sheets[:3]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'=' * 80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'=' * 80}")

        # Show first 25 rows with all columns
        max_rows = min(25, ws.max_row)

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
            # Clean up None values and convert to string safely
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    try:
                        row_data.append(str(cell))
                    except:
                        row_data.append("[?]")

            # Print row with column separators
            print(f"Row {row_idx:2d}: {' | '.join(row_data[:12])}")  # Show first 12 columns

        print(f"\n(Sheet has {ws.max_row} rows and {ws.max_column} columns)")

# Show summary for all team members
print("\n" + "=" * 80)
print("SUMMARY: All Team Member Sheets")
print("=" * 80)
for sheet_name in team_member_sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"{sheet_name:15s}: {ws.max_row:4d} rows x {ws.max_column:2d} columns")

wb.close()
