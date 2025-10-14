import openpyxl
import sys

# Force UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('documents/SCI Workload Tracker - New System.xlsx', data_only=True)

# Check if Dashboard sheet exists
if 'Dashboard' not in wb.sheetnames:
    print("Dashboard sheet not found!")
    wb.close()
    exit(1)

ws = wb['Dashboard']

print("=" * 100)
print("DASHBOARD TAB ANALYSIS")
print("=" * 100)
print(f"Sheet has {ws.max_row} rows and {ws.max_column} columns\n")

# Read all rows to understand structure
print("=" * 100)
print("DASHBOARD CONTENT (showing first 100 rows with data)")
print("=" * 100)

row_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), start=1):
    # Check if row has any data
    has_data = any(cell is not None and str(cell).strip() for cell in row)

    if has_data:
        row_count += 1
        # Clean up None values and convert to string safely
        row_data = []
        for cell in row:
            if cell is None:
                row_data.append("")
            else:
                try:
                    cell_str = str(cell)
                    # Truncate very long cells
                    if len(cell_str) > 50:
                        cell_str = cell_str[:47] + "..."
                    row_data.append(cell_str)
                except:
                    row_data.append("[?]")

        # Print row with column separators (show first 10 columns)
        row_display = ' | '.join(row_data[:10])
        print(f"Row {row_idx:3d}: {row_display}")

print(f"\n(Showing {row_count} rows with data out of first 100 rows)")

# Check for specific patterns - look for column headers
print("\n" + "=" * 100)
print("LOOKING FOR KEY SECTIONS...")
print("=" * 100)

# Scan through rows looking for section headers
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
    first_cell = row[0] if row else None
    if first_cell and str(first_cell).strip():
        first_str = str(first_cell).strip()
        # Look for key terms that might indicate sections
        if any(keyword in first_str.upper() for keyword in ['NAME', 'TOTAL', 'ASSIGNMENT', 'WORKLOAD', 'CAPACITY', 'HOURS', 'SUMMARY']):
            row_data = [str(cell) if cell else "" for cell in row[:15]]
            print(f"\nRow {row_idx}: {' | '.join(row_data)}")

wb.close()
