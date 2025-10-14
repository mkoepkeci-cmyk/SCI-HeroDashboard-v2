import openpyxl
import sys

# Load the workbook
wb = openpyxl.load_workbook('documents/SCI Workload Tracker - New System.xlsx', data_only=True)

# Print all sheet names
print("=" * 80)
print("SHEET NAMES:")
print("=" * 80)
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

print("\n" + "=" * 80)
print("SHEET PREVIEWS:")
print("=" * 80)

# For each sheet, show first 20 rows
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")

    max_rows = min(20, ws.max_row)
    max_cols = min(15, ws.max_column)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), start=1):
        # Clean up None values
        row_data = [str(cell) if cell is not None else "" for cell in row]
        print(f"Row {row_idx:2d}: {' | '.join(row_data[:10])}")  # Show first 10 columns

    print(f"\n(Sheet has {ws.max_row} rows and {ws.max_column} columns)")

wb.close()
