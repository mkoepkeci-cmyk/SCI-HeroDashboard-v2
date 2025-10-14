import openpyxl
import sys

# Force UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('documents/SCI Workload Tracker - New System.xlsx', data_only=True)

# Team member sheets
team_member_sheets = [
    'Ashley', 'Brooke', 'Dawn', 'Jason', 'Josh', 'Kim', 'Lisa',
    'Marisa', 'Marty', 'Matt', 'Melissa', 'Robin', 'Sherry',
    'Trudy', 'Van', 'Yvette', 'Tiffany', 'Carrie'
]

print("=" * 100)
print("TEAM MEMBER ASSIGNMENT ANALYSIS")
print("=" * 100)
print(f"\n{'Name':<12} {'Total':>6} {'Active':>6} {'Complete':>8} {'On Hold':>8} {'Work Types':<50}")
print("-" * 100)

all_data = []

for sheet_name in team_member_sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Count assignments (rows with data, excluding header and empty rows)
        total = 0
        active = 0
        complete = 0
        on_hold = 0
        work_types = {}

        # Read all rows starting from row 2 (after header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Check if this is an assignment row (has SCI name in column A)
            if row[0] and str(row[0]).strip():
                assignment_name = row[1]  # Column B - Assignment
                status = row[8] if len(row) > 8 else None  # Column I - Status
                work_type = row[6] if len(row) > 6 else None  # Column G - Work Type

                # Only count if there's an assignment name
                if assignment_name and str(assignment_name).strip():
                    total += 1

                    # Count by status
                    if status:
                        status_str = str(status).strip()
                        if 'Complete' in status_str:
                            complete += 1
                        elif 'Hold' in status_str:
                            on_hold += 1
                        elif 'Progress' in status_str or 'Planning' in status_str:
                            active += 1

                    # Count by work type
                    if work_type:
                        wt = str(work_type).strip()
                        work_types[wt] = work_types.get(wt, 0) + 1

        # Format work types for display
        wt_display = ", ".join([f"{k}: {v}" for k, v in sorted(work_types.items())])
        if len(wt_display) > 47:
            wt_display = wt_display[:47] + "..."

        print(f"{sheet_name:<12} {total:>6} {active:>6} {complete:>8} {on_hold:>8} {wt_display:<50}")

        all_data.append({
            'name': sheet_name,
            'total': total,
            'active': active,
            'complete': complete,
            'on_hold': on_hold,
            'work_types': work_types
        })

# Print totals
print("-" * 100)
total_all = sum(d['total'] for d in all_data)
active_all = sum(d['active'] for d in all_data)
complete_all = sum(d['complete'] for d in all_data)
on_hold_all = sum(d['on_hold'] for d in all_data)
print(f"{'TOTAL':<12} {total_all:>6} {active_all:>6} {complete_all:>8} {on_hold_all:>8}")

# Print top 10 by total assignments
print("\n" + "=" * 100)
print("TOP 10 TEAM MEMBERS BY TOTAL ASSIGNMENTS")
print("=" * 100)
sorted_data = sorted(all_data, key=lambda x: x['total'], reverse=True)[:10]
for idx, d in enumerate(sorted_data, 1):
    print(f"{idx:2d}. {d['name']:<12} - {d['total']:>3} assignments ({d['active']:>2} active, {d['complete']:>2} complete)")

# Aggregate work types across all team members
print("\n" + "=" * 100)
print("WORK TYPE DISTRIBUTION (ALL TEAM MEMBERS)")
print("=" * 100)
all_work_types = {}
for d in all_data:
    for wt, count in d['work_types'].items():
        all_work_types[wt] = all_work_types.get(wt, 0) + count

for wt, count in sorted(all_work_types.items(), key=lambda x: x[1], reverse=True):
    print(f"  {wt:<30}: {count:>4} assignments")

wb.close()
