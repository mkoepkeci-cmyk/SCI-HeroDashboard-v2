import openpyxl
import sys

# Force UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('documents/SCI Workload Tracker - New System.xlsx', data_only=True)

ws = wb['Dashboard']

print("=" * 120)
print("DASHBOARD TAB - COMPLETE STRUCTURE")
print("=" * 120)

# Get header row (row 1)
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

print("\nCOLUMN HEADERS:")
print("-" * 120)
for idx, header in enumerate(header_row, start=1):
    if header:
        print(f"  Col {idx:2d}: {header}")

# Get all data rows
print("\n" + "=" * 120)
print("TEAM MEMBER DATA:")
print("=" * 120)

data_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=20, values_only=True), start=2):
    # Check if first column (Name) has data
    if row[0]:
        data_rows.append(row)

# Print formatted table
headers = [str(h) if h else "" for h in header_row[:15]]
print(f"\n{headers[0]:<10} {headers[1]:<6} {headers[2]:<8} {headers[3]:<12} {headers[4]:<10} {headers[5]:<12} {headers[6]:<30}")
print("-" * 120)

for row in data_rows:
    name = str(row[0]) if row[0] else ""
    total = f"{row[1]:.0f}" if row[1] else "0"
    active = f"{row[2]:.0f}" if row[2] else "0"
    active_hrs = f"{row[3]:.2f}" if row[3] else "0.00"
    avail_hrs = f"{row[4]:.0f}" if row[4] else "40"
    capacity_util = f"{row[5]*100:.1f}%" if row[5] else "0%"
    capacity_status = str(row[6])[:30] if row[6] else ""

    print(f"{name:<10} {total:>6} {active:>8} {active_hrs:>12} {avail_hrs:>10} {capacity_util:>12} {capacity_status:<30}")

# Show work type breakdowns
print("\n" + "=" * 120)
print("WORK TYPE BREAKDOWN (by hours):")
print("=" * 120)

# Find work type columns (they seem to be after column 7)
print(f"\n{'Name':<10} {'Epic Gold':<15} {'Governance':<15} {'Sys Init':<15} {'Sys Proj':<15}")
print("-" * 80)

for row in data_rows:
    name = str(row[0]) if row[0] else ""

    # Columns 7-8: EpicGold
    epic_gold = f"{row[7]:.0f} cnt, {row[8]:.1f}h" if row[7] and row[8] else "0 cnt, 0h"

    # Columns 9-10: Governance
    governance = f"{row[9]:.0f} cnt, {row[10]:.1f}h" if row[9] and row[10] else "0 cnt, 0h"

    # Columns 11-12: System Initiative
    sys_init = f"{row[11]:.0f} cnt, {row[12]:.1f}h" if row[11] and row[12] else "0 cnt, 0h"

    # Columns 13-14: System Project
    sys_proj = f"{row[13]:.0f} cnt, {row[14]:.1f}h" if row[13] and row[14] else "0 cnt, 0h"

    print(f"{name:<10} {epic_gold:<15} {governance:<15} {sys_init:<15} {sys_proj:<15}")

# Calculate totals
print("\n" + "=" * 120)
print("TEAM SUMMARY:")
print("=" * 120)

total_assignments = sum(row[1] for row in data_rows if row[1])
total_active = sum(row[2] for row in data_rows if row[2])
total_hours = sum(row[3] for row in data_rows if row[3])
avg_capacity = sum(row[5] for row in data_rows if row[5]) / len(data_rows) if data_rows else 0

print(f"Total Assignments: {total_assignments:.0f}")
print(f"Total Active: {total_active:.0f}")
print(f"Total Active Hours/Week: {total_hours:.2f}")
print(f"Average Capacity Utilization: {avg_capacity*100:.1f}%")

# Check for over/near capacity
over_capacity = [row[0] for row in data_rows if row[5] and row[5] > 1.0]
near_capacity = [row[0] for row in data_rows if row[5] and 0.8 <= row[5] <= 1.0]
under_capacity = [row[0] for row in data_rows if row[5] and row[5] < 0.8]

print(f"\nOver Capacity (>100%): {len(over_capacity)} - {', '.join(over_capacity)}")
print(f"Near Capacity (80-100%): {len(near_capacity)} - {', '.join(near_capacity)}")
print(f"Under Capacity (<80%): {len(under_capacity)} - {', '.join(under_capacity)}")

wb.close()
