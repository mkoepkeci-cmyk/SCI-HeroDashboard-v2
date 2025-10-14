import openpyxl
import sys

# Force UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('documents/SCI Workload Tracker - New System.xlsx', data_only=True)

ws = wb['Workload']

print("=" * 120)
print("WORKLOAD TAB ANALYSIS")
print("=" * 120)
print(f"Sheet has {ws.max_row} rows and {ws.max_column} columns\n")

# Get header row (row 1)
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

print("COLUMN HEADERS:")
print("-" * 120)
for idx, header in enumerate(header_row[:15], start=1):
    if header:
        print(f"  Col {idx:2d}: {header}")

# Get data rows (first 20)
print("\n" + "=" * 120)
print("WORKLOAD DATA (First 20 rows):")
print("=" * 120)

for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=21, values_only=True), start=1):
    # Check if row has any data
    has_data = any(cell is not None and str(cell).strip() for cell in row[:10])

    if has_data:
        row_data = []
        for cell in row[:10]:
            if cell is None:
                row_data.append("")
            else:
                try:
                    cell_str = str(cell)
                    if len(cell_str) > 30:
                        cell_str = cell_str[:27] + "..."
                    row_data.append(cell_str)
                except:
                    row_data.append("[?]")

        print(f"Row {row_idx:3d}: {' | '.join(row_data)}")

wb.close()
